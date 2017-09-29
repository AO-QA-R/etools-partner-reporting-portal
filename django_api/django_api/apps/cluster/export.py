from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment

from django.conf import settings
from django.db.models import Count

import itertools

from indicator.models import Disaggregation, DisaggregationValue

PATH = settings.BASE_DIR + "/apps/cluster/templates/excel/export.xlsx"
SAVE_PATH = settings.MEDIA_ROOT + '/'

class XLSXWriter:
    def __init__(self, indicators, response_plan_id):

        self.wb = load_workbook(PATH)
        self.sheet = self.wb.get_active_sheet()
        self.indicators = indicators
        self.response_plan_id = response_plan_id


    def duplicate_sheet(self, sheet):
        return self.wb.copy_worksheet(sheet)


    def generate_sheet(self, disaggregation_types):

        DISAGGREGATION_COLUMN_START = 29
        INDICATOR_DATA_ROW_START = 4

        # Setup a title
        self.sheet.title = ",".join([dt.name for dt in disaggregation_types]) if disaggregation_types else "None"

        # Generate disaggregation types columns
        disaggregation_types_map = dict() # it holds column number for given Disaggregation Type ID
        for idx, dt in enumerate(disaggregation_types):
            # Name
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + idx).value = dt.name
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + idx).alignment = Alignment(horizontal='center')
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + idx).font = Font(bold=True)
            # ID
            self.sheet.cell(row=2, column=DISAGGREGATION_COLUMN_START + idx).value = dt.id
            self.sheet.cell(row=2, column=DISAGGREGATION_COLUMN_START + idx).alignment = Alignment(horizontal='center')
            disaggregation_types_map[dt.id] = DISAGGREGATION_COLUMN_START + idx


        # Prepare Disaggregation Values for given disaggregation types
        disaggregation_values = list()
        disaggregation_values_list = list()
        disaggregation_values_map = dict()
        for disaggregation_type in disaggregation_types:
                disaggregation_values_base = DisaggregationValue.objects.filter(disaggregation=disaggregation_type).order_by('value')
                disaggregation_values_list.append(disaggregation_values_base)

        # Create all possible combinations (max. 3 items per combination) for disaggregation values
        for combination_items in range(1, 4):
            for combinations in itertools.combinations(disaggregation_values_list, combination_items):
                disaggregation_values += list(itertools.product(*combinations))

        # Generate disaggregation values columns
        for idx, dvs in enumerate(disaggregation_values):
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).value =  " + ".join([dv.value for dv in dvs])
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).alignment = Alignment(horizontal='center')
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).font = Font(bold=True)
            self.sheet.cell(row=2, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).value = ", ".join([str(dv.id) for dv in dvs])
            disaggregation_values_map[", ".join([str(dv.id) for dv in dvs])] = DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx

        # Generate Total disaggregation value
        self.sheet.cell(row=1,column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + len(disaggregation_values)).value = "Total"
        self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + len(disaggregation_values)).alignment = Alignment(horizontal='center')
        disaggregation_values_map["()"] = DISAGGREGATION_COLUMN_START + len(disaggregation_types) + len(disaggregation_values)

        indicators = self.indicators
        if disaggregation_types:
            indicators = indicators.annotate(count=Count('reportable__disaggregations')).filter(count=len(disaggregation_types))
            for dt in disaggregation_types:
                indicators = indicators.filter(reportable__disaggregations=dt)
        else:
            indicators = indicators.filter(reportable__disaggregations__isnull=True)

        indicators = indicators.order_by('reportable__id')

        if not indicators:
            return False

        # Each row is one location per indicator
        start_row_id = INDICATOR_DATA_ROW_START

        for indicator in indicators:
            cluster_objective = indicator.reportable.cluster_objectives.first()
            cluster_activity = indicator.reportable.cluster_activities.first()
            partner_project = indicator.reportable.partner_projects.first()
            partner_activity = indicator.reportable.partner_activities.first()
            locations_data = indicator.indicator_location_data

            cluster = None

            if cluster_objective:
                cluster = cluster_objective.cluster
                partner_activity = cluster_objective.cluster_activities.first().partner_activities.first()
                partner_project = cluster_objective.cluster.partner_projects.first()
            elif cluster_activity:
                cluster = cluster_activity.cluster_objective.cluster
                cluster_objective = cluster_activity.cluster_objective
                partner_activity = cluster_activity.partner_activities.first()
                partner_project = cluster_activity.cluster_objective.cluster.partner_projects.first()
            elif partner_activity:
                cluster = partner_activity.cluster_activity.cluster_objective.cluster
                cluster_objective = partner_activity.cluster_activity.cluster_objective
                partner_project = partner_activity.project
            elif partner_project:
                cluster = partner_project.clusters.first()
                cluster_objective = partner_project.clusters.first().cluster_objectives.first()
                partner_activity = partner_project.partner.partner_activities.first()

            for location_data in locations_data.all():

                self.sheet.cell(row=start_row_id, column=1).value = cluster.get_type_display()
                self.sheet.cell(row=start_row_id, column=2).value = cluster.partner_projects.first().partner.title
                self.sheet.cell(row=start_row_id, column=3).value = cluster_objective.title
                self.sheet.cell(row=start_row_id, column=4).value = partner_activity.title
                self.sheet.cell(row=start_row_id, column=5).value = indicator.reportable.blueprint.title
                self.sheet.cell(row=start_row_id, column=6).value = partner_project.title
                self.sheet.cell(row=start_row_id, column=7).value = partner_project.get_status_display()
                self.sheet.cell(row=start_row_id, column=8).value = partner_activity.start_date
                self.sheet.cell(row=start_row_id, column=9).value = partner_activity.end_date

                self.sheet.cell(row=start_row_id, column=10).value = indicator.reportable.baseline
                self.sheet.cell(row=start_row_id, column=11).value = indicator.reportable.target
                self.sheet.cell(row=start_row_id, column=12).value = indicator.reportable.total['c']

                self.sheet.cell(row=start_row_id, column=13).value =location_data.location.title
                self.sheet.cell(row=start_row_id, column=14).value = location_data.location.p_code
                self.sheet.cell(row=start_row_id, column=15).value = "XXX" # TODO: no field here!

                self.sheet.cell(row=start_row_id, column=16).value = indicator.time_period_start
                self.sheet.cell(row=start_row_id, column=17).value = indicator.time_period_end

                self.sheet.cell(row=start_row_id, column=18).value = indicator.get_report_status_display()
                self.sheet.cell(row=start_row_id, column=19).value = indicator.get_overall_status_display()

                self.sheet.cell(row=start_row_id, column=20).value = indicator.submission_date
                self.sheet.cell(row=start_row_id, column=21).value = cluster.id
                self.sheet.cell(row=start_row_id, column=22).value = cluster_objective.id
                self.sheet.cell(row=start_row_id, column=23).value = partner_activity.id
                self.sheet.cell(row=start_row_id, column=24).value = indicator.reportable.blueprint.id
                self.sheet.cell(row=start_row_id, column=25).value = partner_project.partner.id
                self.sheet.cell(row=start_row_id, column=26).value = partner_project.id
                self.sheet.cell(row=start_row_id, column=27).value = indicator.id
                self.sheet.cell(row=start_row_id, column=28).value = location_data.id

                # Check location item disaggregation type
                for reported_disaggregation_type in location_data.disaggregation_reported_on:
                    if reported_disaggregation_type in disaggregation_types_map:
                        self.sheet.cell(row=start_row_id, column=disaggregation_types_map[reported_disaggregation_type]).value = "X"

                # Check location item values
                for k, v in location_data.disaggregation.items():
                    if k == "()":
                        self.sheet.cell(row=start_row_id, column=disaggregation_values_map['()']).value = v['c']
                    else:
                        for dk, dv in disaggregation_values_map.items():
                            if dk == "()":
                                continue
                            if sorted(list(eval(k))) == sorted(list(int(k) for k in dk.split(","))):
                                self.sheet.cell(row=start_row_id, column=dv).value = v['c']

                start_row_id += 1

        return True

    def export_data(self):

        MOVE_COLUMN = 0
        MAXIMUM_DISSAGREGATIONS_PER_INDICATOR = 3


        # Disaggregation types
        disaggregation_types_base = Disaggregation.objects.filter(response_plan__id=self.response_plan_id)
        disaggregation_types_base_length = len(disaggregation_types_base)

        # Split data into spreadsheets, combination of disaggregation_types_length with r-length tuples and no repeated elements
        disaggregation_types_list = list()
        for i in range(MAXIMUM_DISSAGREGATIONS_PER_INDICATOR + 1):
            disaggregation_types_list += list(itertools.combinations(disaggregation_types_base, i))

        # Duplicate spreadsheets base
        sheets = [self.sheet, ]
        for i in range(1, len(disaggregation_types_list)):
            sheets.append(self.duplicate_sheet(self.sheet))

        # Generate data per spreadsheet
        sheet_no = 0
        for disaggregation_types in disaggregation_types_list:
            self.sheet = sheets[sheet_no]
            if not self.generate_sheet(disaggregation_types):
                self.wb.remove_sheet(self.sheet)
            sheet_no += 1

        filepath = SAVE_PATH + 'export.xlsx'
        self.wb.save(filepath)
        return filepath